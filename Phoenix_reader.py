#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Dec 24 14:13:20 2022

@author: thomasstinglhamber
"""
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 21 10:59:59 2022

@author: thomasstinglhamber
"""
import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


df = pd.read_excel('/Users/thomasstinglhamber/Desktop/SpotAnalysisResults_2022-12-09.xlsx', header=6)

Angle = 180
Energy = 100
Mu = 1
correction_vector = [0.03,-1.01]



#---------------------------------------------------------

X_nominal=[150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150]
Y_nominal =[150,150,150,150,150,150,90,90,90,90,90,90,30,30,30,30,30,30,-30,-30,-30,-30,-30,-30,-90,-90,-90,-90,-90,-90,-150,-150,-150,-150,-150,-150]

    

# Je tri par nom pour separer les 5 images
name=[]
for i in range(0,len(df.iloc[:, [0]])):
    if df.iloc[i, [0]].to_string(index=False) not in name:
        name.append(df.iloc[i, [0]].to_string(index=False))


# je met tout ces tableaux dans une liste "merge"
merge=[]
grouped = df.groupby(df['Image Name'])
for k in name:
    df_new = grouped.get_group(k)
    merge.append(df_new)

#Je tri tout ces tableaux selon Y
for i in range(0,len(merge)):
    merge[i].sort_values(['Y (mm)','X (mm)'], 
               ascending=[False,False], inplace=True,kind='mergesort')

#Je tri tout ces tableaux selon X
for i in range(0,len(merge)):
    for j in [0,1,2,3,4,5]:
        merge[i].iloc[0+6*j:6+6*j, [1,2]]= merge[i].iloc[0+6*j:6+6*j, [1,2]].sort_values(['X (mm)'], 
                   ascending=[False])
    #print(merge[i]) 
    merge[i]=merge[i].reset_index(drop = True)
      
    #print(merge[i])   
#print(merge)



#je met tout les colonnes X et Y dans un tableau specifique pour calculer les mean et std
df_stat_X= pd.DataFrame(merge[0].iloc[:, [1]])+correction_vector[0]
df_stat_X['tab1']=merge[1].iloc[:, [1]]+correction_vector[0]
df_stat_X['tab2']=merge[2].iloc[:, [1]]+correction_vector[0]
df_stat_X['tab3']=merge[3].iloc[:, [1]]+correction_vector[0]
df_stat_X['tab4']=merge[4].iloc[:, [1]]+correction_vector[0]
#print(df_stat_X)
# j'enleve la position nominal X
df_stat_X['X (mm)']=X_nominal-df_stat_X['X (mm)']
df_stat_X['tab1']=X_nominal-df_stat_X['tab1']
df_stat_X['tab2']=X_nominal-df_stat_X['tab2']
df_stat_X['tab3']=X_nominal-df_stat_X['tab3']
df_stat_X['tab4']=X_nominal-df_stat_X['tab4']
#print(df_stat_X)


# Ici tableau selon Y
df_stat_Y= pd.DataFrame(merge[0].iloc[:, [2]])+correction_vector[1]
df_stat_Y['tab1']=merge[1].iloc[:, [2]]+correction_vector[1]
df_stat_Y['tab2']=merge[2].iloc[:, [2]]+correction_vector[1]
df_stat_Y['tab3']=merge[3].iloc[:, [2]]+correction_vector[1]
df_stat_Y['tab4']=merge[4].iloc[:, [2]]+correction_vector[1]
#print(df_stat_Y)
# j'enleve la position nominal Y
df_stat_Y['Y (mm)']=Y_nominal-df_stat_Y['Y (mm)']
df_stat_Y['tab1']=Y_nominal-df_stat_Y['tab1']
df_stat_Y['tab2']=Y_nominal-df_stat_Y['tab2']
df_stat_Y['tab3']=Y_nominal-df_stat_Y['tab3']
df_stat_Y['tab4']=Y_nominal-df_stat_Y['tab4']
#print(df_stat_Y)

# =============================================================================
# #creation du plot pour vérifier les points et la mean/std
# plt.figure()
# plt.scatter(df_stat_X.mean(axis=1),df_stat_Y.mean(axis=1),c='red',marker='x')
# #plt.scatter(df['X 2D Fit (mm)'],df['Y 2D Fit (mm)'],marker='.')
# # =============================================================================
# # for l in range(0,len(merge)):
# #     plt.scatter(merge[l]['X (mm)']+correction_vector[0],merge[l]['Y (mm)']+correction_vector[1],marker='.',label=merge[l].iloc[1,[0]].to_string(index=False))
# # plt.legend()
# # =============================================================================
# 
# plt.show()
# 
# =============================================================================

# implementer un moyen de stocker les mean/std dans un fichier externe
# faire attention a la nomenclature !

final=pd.DataFrame()


final['Angle']= [Angle]*36
final['Mu']= [Mu]*36
final['Energy']= [Energy]*36

final['Moyenne_X']= df_stat_X.mean(axis=1)
final['Moyenne_Y']= df_stat_Y.mean(axis=1)

final['Std_X']= df_stat_X.std(axis=1)
final['Std_Y']= df_stat_Y.std(axis=1)

#print(final)

#Si on veut trier l'excel de X
# =============================================================================
# df_stat_X.sort_values(['Moyenne_X'],ascending=[False], inplace=True,kind='mergesort')
# df_stat_X=df_stat_X.reset_index(drop = True)
# =============================================================================

final.to_excel('/Users/thomasstinglhamber/Desktop/dataX.xlsx',index=False)



# =============================================================================
# book = load_workbook('/Users/thomasstinglhamber/Desktop/dataX.xlsx')
# writer = pd.ExcelWriter('/Users/thomasstinglhamber/Desktop/dataX.xlsx', engine='openpyxl')
# writer.book = book
# writer.sheets = {ws.title: ws for ws in book.worksheets}
# 
# for sheetname in writer.sheets:
#     final.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
# 
# writer.save()
# 
# =============================================================================
