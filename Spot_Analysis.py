#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 20 18:53:50 2023

@author: thomasstinglhamber
"""

import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import matplotlib.colors as mcolors
import math

creat=0

df1 = pd.read_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New2/Erreurexp_3.xlsx', header=0)

plt.scatter(df1['Moyenne_X'],df1['Moyenne_Y'],color='r')


# Iterate over each row in the data frame
for index, row in df1.iterrows():
    # Get the rotation angle for this row
    angle_degrees1 = row['Angle']
    
    if angle_degrees1 == 0:
        angle_degrees = 0.11
    if angle_degrees1 == 45:
        angle_degrees = -0.04
    if angle_degrees1 == 90:
        angle_degrees = 0.18
    if angle_degrees1 == 180:
        angle_degrees = 0.28
    if angle_degrees1 == 270:
        angle_degrees = 0.16
        
    #print(angle_degrees)
    # Convert the angle from degrees to radians
    angle_radians = math.radians(angle_degrees)

    # Define the rotation matrix
    rotation_matrix = np.array([
        [np.cos(angle_radians), -np.sin(angle_radians)],
        [np.sin(angle_radians), np.cos(angle_radians)]
    ])

    # Get the x and y coordinates for this row
    x = row['Moyenne_X']+row['Nominal_X']
    y = row['Moyenne_Y']+row['Nominal_Y']
    
    # Apply the rotation matrix to the point
    rotated_point = np.dot(rotation_matrix, [x, y])

    # Update the 'Moyenne_X' and 'Moyenne_Y' columns for this row
    df1.at[index, 'Moyenne_X'] = rotated_point[0]-row['Nominal_X']
    df1.at[index, 'Moyenne_Y'] = rotated_point[1]-row['Nominal_Y']

# Update the 'Moyenne_X' and 'Moyenne_Y' columns with the rotated points
#df1[['Moyenne_X', 'Moyenne_Y']] = rotated_points


# Plot the rotated points

plt.scatter(df1['Moyenne_X'], df1['Moyenne_Y'])
plt.show()





if creat>0:
    book = load_workbook('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New2/Erreurexp_3_Rot.xlsx')
    writer = pd.ExcelWriter('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New2/Erreurexp_3_Rot.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    
    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    
    writer.save()

if creat==0:
    df1.to_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New2/Erreurexp_3_Rot.xlsx',index=False)
    creat=creat+1




