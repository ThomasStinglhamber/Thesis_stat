#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jan  1 16:24:44 2023

@author: thomasstinglhamber
"""
import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# Get the list of all files and directories
path = '/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Phoenix/Mesure'
dir_list = os.listdir(path)
merge=[]
creat=0
compt=1
for x in dir_list:
    print(compt,'/',150)
    compt=compt+1
    if x.endswith(".xlsx"):
        # Prints only text file present in My Folder
        Angle = int(x[0:3])
        Energy= int(x[3:6])
        MU= int(x[6:10])/1000
        #print('Angle :',Angle,'Energy :',Energy,'MU :',MU)
        
        df = pd.read_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Phoenix/Mesure/'+str(x), header=6)
        
        if Angle  == 0:
            correction_vector = [0.13,-1.06]
        if Angle  == 45:
            correction_vector = [0.02,-1.53]
        if Angle  == 270:
            correction_vector = [-0.01,-0.23]
        if Angle  == 180:
            correction_vector = [-0.07,-0.20]
        if Angle  == 90:
            correction_vector = [1.02,-0.44]

        #---------------------------------------------------------
        
        X_nominal=[150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150]
        Y_nominal =[150,150,150,150,150,150,90,90,90,90,90,90,30,30,30,30,30,30,-30,-30,-30,-30,-30,-30,-90,-90,-90,-90,-90,-90,-150,-150,-150,-150,-150,-150]
        
        
        tri1= [150,90,30,-30,-90,-150]
        
        # Je tri par nom pour separer les 5 images
        name=[]
        for i in range(0,len(df.iloc[:, [0]])):
            if df.iloc[i, [0]].to_string(index=False) not in name:
                name.append(df.iloc[i, [0]].to_string(index=False))
        
        
        # je met tout ces tableaux dans une liste "merge"
        merge=[]
        merge2=[]
        grouped = df.groupby(df['Image Name'])
        for k in name:
            df_new = grouped.get_group(k)
            df_FINAL=pd.DataFrame()
        
            if len(df_new)>=36:  # check si MyQA a mal interpreter l'image
                
                df_new.sort_values(['Y (mm)','X (mm)'],
                           ascending=[False,False], inplace=True,kind='mergesort')
               
                number=[]
                sort1= [0,0,0,0,0,0]
                sort2= [0,0,0,0,0,0]
                
                for k in range(len(tri1)): 
                    for j in range(len(df_new.iloc[:,[2]])):
                        # je tri en fonction d'un interval
                        if tri1[k]-30 <= float(df_new.iloc[j,[2]].to_string(index=False)) < tri1[k]+30:
                            sort1[k]=sort1[k]+1
                                            
                somme=0
                
                for j in sort1:
                    df_new.iloc[somme:somme+j, [1,2]]= df_new.iloc[somme:somme+j, [1,2]].sort_values(['X (mm)'], 
                               ascending=[False])
                    
                    df_transi= pd.DataFrame(df_new.iloc[somme:somme+j, [1,2,3,4]])
                    if j != 6:
                        
                        sort2= [0,0,0,0,0,0]
                        for m in range(len(tri1)):
                            for r in range(len(df_transi.iloc[:,[0]])):
                                
                                if tri1[m]-30 <= float(df_transi.iloc[r,[0]].to_string(index=False)) < tri1[m]+30:
                                    sort2[m]=sort2[m]+1
                                    
                        somme2=0
                        for a,y in enumerate(sort2):
                            df_transi.iloc[a,[2,3]]=0
                            if y > 1:
                                df_transi.iloc[a,[2,3]] = (df_transi.iloc[somme2:somme2+y,[0,1]]-tri1[a]).std()
                            df_transi.iloc[a,[0,1]] = (df_transi.iloc[somme2:somme2+y,[0,1]]).mean()
                            somme2=somme2+y
                            
                        df_transi = df_transi.drop(df_transi.index[6:])
                        df_FINAL = df_FINAL.append(df_transi, ignore_index=True)
                    else : 
                        df_transi.iloc[:, [2,3]] = 0
                        
                        df_FINAL = df_FINAL.append(df_transi, ignore_index=True)
                    
                    
                    somme = somme +j
                
            
                merge.append(df_FINAL)
                merge2=merge
            else : 
                merge.append(df_new)
                #Je tri tout ces tableaux selon Y
                for i in range(0,len(merge)):
                    merge[i].sort_values(['Y (mm)','X (mm)'], 
                               ascending=[False,False], inplace=True,kind='mergesort')
                
                #Je tri tout ces tableaux selon X
                for i in range(0,len(merge)):
                    for j in [0,1,2,3,4,5]:
                        merge[i].iloc[0+6*j:6+6*j, [1,2]]= merge[i].iloc[0+6*j:6+6*j, [1,2]].sort_values(['X (mm)'], 
                                   ascending=[False])
                    
                    merge[i]=merge[i].reset_index(drop = True)
                    merge[i].iloc[:, [3,4]] = 0
                    merge2.append(merge[i].iloc[:, [1,2,3,4]])
                    
        
        #merge[0].to_excel('/Users/thomasstinglhamber/Desktop/TESTT.xlsx',index=False)
        
        
        #je met tout les colonnes X et Y dans un tableau specifique pour calculer les mean et std
        df_stat_Xmean= pd.DataFrame(merge2[0].iloc[:, [0]])+correction_vector[0]
        df_stat_Xmean['tab1']=merge2[1].iloc[:, [0]]+correction_vector[0]
        df_stat_Xmean['tab2']=merge2[2].iloc[:, [0]]+correction_vector[0]
        df_stat_Xmean['tab3']=merge2[3].iloc[:, [0]]+correction_vector[0]
        df_stat_Xmean['tab4']=merge2[4].iloc[:, [0]]+correction_vector[0]
        #print(df_stat_X)
        
# =============================================================================
#         # j'enleve la position nominal X
#         df_stat_Xmean['X (mm)']=X_nominal-df_stat_Xmean['X (mm)']
#         df_stat_Xmean['tab1']=X_nominal-df_stat_Xmean['tab1']
#         df_stat_Xmean['tab2']=X_nominal-df_stat_Xmean['tab2']
#         df_stat_Xmean['tab3']=X_nominal-df_stat_Xmean['tab3']
#         df_stat_Xmean['tab4']=X_nominal-df_stat_Xmean['tab4']
#         
# =============================================================================
        #df_stat_Xmean.to_excel('/Users/thomasstinglhamber/Desktop/dataXmean.xlsx',index=False)
        
        # Ici tableau selon Y
        df_stat_Ymean= pd.DataFrame(merge2[0].iloc[:, [1]])+correction_vector[1]
        df_stat_Ymean['tab1']=merge2[1].iloc[:, [1]]+correction_vector[1]
        df_stat_Ymean['tab2']=merge2[2].iloc[:, [1]]+correction_vector[1]
        df_stat_Ymean['tab3']=merge2[3].iloc[:, [1]]+correction_vector[1]
        df_stat_Ymean['tab4']=merge2[4].iloc[:, [1]]+correction_vector[1]
        
# =============================================================================
#         # j'enleve la position nominal Y
#         df_stat_Ymean['Y (mm)']=Y_nominal-df_stat_Ymean['Y (mm)']
#         df_stat_Ymean['tab1']=Y_nominal-df_stat_Ymean['tab1']
#         df_stat_Ymean['tab2']=Y_nominal-df_stat_Ymean['tab2']
#         df_stat_Ymean['tab3']=Y_nominal-df_stat_Ymean['tab3']
#         df_stat_Ymean['tab4']=Y_nominal-df_stat_Ymean['tab4']
#         #print(df_stat_Y)
# =============================================================================
        
        
        #df_stat_Ymean.to_excel('/Users/thomasstinglhamber/Desktop/dataYmean.xlsx',index=False)
        
        df_stat_Xstd= pd.DataFrame(merge2[0].iloc[:, [2]])
        df_stat_Xstd['tab1']=merge2[1].iloc[:, [2]]
        df_stat_Xstd['tab2']=merge2[2].iloc[:, [2]]
        df_stat_Xstd['tab3']=merge2[3].iloc[:, [2]]
        df_stat_Xstd['tab4']=merge2[4].iloc[:, [2]]
        
        df_stat_Xstd['sommequadX']=(np.sqrt(merge2[0].iloc[:, [2]]**2 +merge2[1].iloc[:, [2]]**2 +merge2[2].iloc[:, [2]]**2 +merge2[3].iloc[:, [2]]**2 +merge2[4].iloc[:, [2]]**2))
        
        #df_stat_Xstd.to_excel('/Users/thomasstinglhamber/Desktop/dataXstd.xlsx',index=False)
        
        
        
        df_stat_Ystd= pd.DataFrame(merge2[0].iloc[:, [3]])
        df_stat_Ystd['tab1']=merge2[1].iloc[:, [3]]
        df_stat_Ystd['tab2']=merge2[2].iloc[:, [3]]
        df_stat_Ystd['tab3']=merge2[3].iloc[:, [3]]
        df_stat_Ystd['tab4']=merge2[4].iloc[:, [3]]
        
        df_stat_Ystd['sommequadY']=(np.sqrt(merge2[0].iloc[:, [3]]**2 +merge2[1].iloc[:, [3]]**2 +merge2[2].iloc[:, [3]]**2 +merge2[3].iloc[:, [3]]**2 +merge2[4].iloc[:, [3]]**2))
        
        #df_stat_Ystd.to_excel('/Users/thomasstinglhamber/Desktop/dataYstd.xlsx',index=False)
        
        final=pd.DataFrame()
        
        
        final['Angle']= [Angle]*36
        final['Mu']= [MU]*36
        final['Energy']= [Energy]*36
        
        final['Moyenne_X']= df_stat_Xmean.mean(axis=1)
        final['Moyenne_Y']= df_stat_Ymean.mean(axis=1)
        
        final['Std_X']= np.sqrt(df_stat_Xmean.std(axis=1)**2+df_stat_Xstd['sommequadX']**2)
        final['Std_Y']= np.sqrt(df_stat_Ymean.std(axis=1)**2+df_stat_Ystd['sommequadY']**2)
        
        
        #Si on veut trier l'excel de X
        # =============================================================================
        # df_stat_X.sort_values(['Moyenne_X'],ascending=[False], inplace=True,kind='mergesort')
        # df_stat_X=df_stat_X.reset_index(drop = True)
        # =============================================================================
        
        if creat>0:
            book = load_workbook('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Phoenix/Merge.xlsx')
            writer = pd.ExcelWriter('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Phoenix/Merge.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            
            for sheetname in writer.sheets:
                final.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
            
            writer.save()
        
        if creat==0:
            final.to_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Phoenix/Merge.xlsx',index=False)
            creat=creat+1
        
    
