#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jan 23 14:46:29 2023

@author: thomasstinglhamber
"""
import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook




# Get the list of all files and directories
path = '/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Log_rename'
dir_list = os.listdir(path)
merge=[]
merge2=[]
creat=0
tri1= [150,90,30,-30,-90,-150]
X_nominal=[150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150,150,90,30,-30,-90,-150]
Y_nominal =[150,150,150,150,150,150,90,90,90,90,90,90,30,30,30,30,30,30,-30,-30,-30,-30,-30,-30,-90,-90,-90,-90,-90,-90,-150,-150,-150,-150,-150,-150]


y=1
for x in dir_list:
    #x2=sorted(x)
    #print(x)
    if x.endswith(".csv"):
        # Prints only text file present in My Folder
        numbers = list(range(151, 181)) + list(range(331, 356))+ list(range(806, 866))
        #print(numbers)
        #print(y)
        #merge.append(pd.read_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Mesure/Phoenix/Raw_excel/'+str(x),header=6))
        if y not in numbers:
            #print(y)
            merge.append(pd.read_csv("/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Log_rename/Log_"+str("{:04}".format(y))+'.csv', delimiter=','))#, skiprows=[0])
        
        y=y+1

#print(len(merge))
        if len(merge)==5:
            for k in range(len(merge)):
                df_new = merge[k]
                #print(df_new)
                
                
                
                df_FINAL=pd.DataFrame()
        
                if len(df_new)>=36:  # check si MyQA a mal interpreter l'image
                
                    df_new.sort_values(['y','x'],
                                   ascending=[False,False], inplace=True,kind='mergesort')


                    #print(df_new)
                    number=[]
                    sort1= [0,0,0,0,0,0]
                    sort2= [0,0,0,0,0,0]
                    
                    for k in range(len(tri1)): 
                        for j in range(len(df_new.iloc[:,[2]])):
                            # je tri en fonction d'un interval
                            if tri1[k]-30 <= float(df_new.iloc[j,[7]].to_string(index=False)) < tri1[k]+30:
                                sort1[k]=sort1[k]+1
                                                
                    somme=0
                    #print(sort1)
                    
                    for j in sort1:
                        df_new.iloc[somme:somme+j, [6,7]]= df_new.iloc[somme:somme+j, [6,7]].sort_values(['x'], 
                                   ascending=[False])
                        
                        df_transi= pd.DataFrame(df_new.iloc[somme:somme+j, [3,5,6,7,8]])
                        #print(df_transi)
                        
                        df_FINAL = df_FINAL.append(df_transi, ignore_index=True)
                        
                        
                        somme = somme +j
                        
                    df_FINAL = df_FINAL[['gantry_angle', 'mu', 'energy', 'x','y']]


                    df_FINAL = df_FINAL.rename(columns={'gantry_angle': 'Angle', 'mu': 'Mu', 'energy': 'Energy', 'x': 'X','y':'Y'})

                    merge2.append(df_FINAL)
                    #merge2=merge
            
            #print(merge2)
            
            df_stat_Xmean= pd.DataFrame(merge2[0].iloc[:, [3]])
            df_stat_Xmean['tab1']=merge2[1].iloc[:, [3]]
            df_stat_Xmean['tab2']=merge2[2].iloc[:, [3]]
            df_stat_Xmean['tab3']=merge2[3].iloc[:, [3]]
            df_stat_Xmean['tab4']=merge2[4].iloc[:, [3]]
            #print(df_stat_Xmean)
# =============================================================================
#             # j'enleve la position nominal X
#             df_stat_Xmean['X']=X_nominal-df_stat_Xmean['X']
#             df_stat_Xmean['tab1']=X_nominal-df_stat_Xmean['tab1']
#             df_stat_Xmean['tab2']=X_nominal-df_stat_Xmean['tab2']
#             df_stat_Xmean['tab3']=X_nominal-df_stat_Xmean['tab3']
#             df_stat_Xmean['tab4']=X_nominal-df_stat_Xmean['tab4']
#             #print(df_stat_Xmean)
# =============================================================================
                
            # Ici tableau selon Y
            df_stat_Ymean= pd.DataFrame(merge2[0].iloc[:, [4]])
            df_stat_Ymean['tab1']=merge2[1].iloc[:, [4]]
            df_stat_Ymean['tab2']=merge2[2].iloc[:, [4]]
            df_stat_Ymean['tab3']=merge2[3].iloc[:, [4]]
            df_stat_Ymean['tab4']=merge2[4].iloc[:, [4]]
            
# =============================================================================
#             # j'enleve la position nominal Y
#             df_stat_Ymean['Y']=Y_nominal-df_stat_Ymean['Y']
#             df_stat_Ymean['tab1']=Y_nominal-df_stat_Ymean['tab1']
#             df_stat_Ymean['tab2']=Y_nominal-df_stat_Ymean['tab2']
#             df_stat_Ymean['tab3']=Y_nominal-df_stat_Ymean['tab3']
#             df_stat_Ymean['tab4']=Y_nominal-df_stat_Ymean['tab4']
#             #print(df_stat_Ymean)
# =============================================================================
                
            final=pd.DataFrame()
            
            
            #print(merge2[1].iloc[:, [2]])
            if 190<float(merge2[1].iloc[0, [2]].to_string(index=False))< 210:
                Energy=200
            if 60<float(merge2[1].iloc[0, [2]].to_string(index=False))< 80:
                Energy=70
            if 90<float(merge2[1].iloc[0, [2]].to_string(index=False))< 110:
                Energy=100
            if 140<float(merge2[1].iloc[0, [2]].to_string(index=False))< 160:
                Energy=150
            if 216<float(merge2[1].iloc[0, [2]].to_string(index=False))< 236:
                Energy=226
            
            if -10<float(merge2[1].iloc[0, [0]].to_string(index=False))< 10:
                Angle=0
            if 35<float(merge2[1].iloc[0, [0]].to_string(index=False))< 55:
                Angle=45
            if 260<float(merge2[1].iloc[0, [0]].to_string(index=False))< 280:
                Angle=270
            if 170<float(merge2[1].iloc[0, [0]].to_string(index=False))< 190:
                Angle=180
            if 80<float(merge2[1].iloc[0, [0]].to_string(index=False))< 100:
                Angle=90
            
            if 0<float(merge2[1].iloc[0, [1]].to_string(index=False))< 0.030:
                MU=0.015
            if 0.05<float(merge2[1].iloc[0, [1]].to_string(index=False))< 0.2:
                MU=0.1
            if 0.3<float(merge2[1].iloc[0, [1]].to_string(index=False))< 0.7:
                MU=0.5
            if 0.8<float(merge2[1].iloc[0, [1]].to_string(index=False))< 1.2:
                MU=1
            if 1.8<float(merge2[1].iloc[0, [1]].to_string(index=False))< 2.2:
                MU=2
            if 4.5<float(merge2[1].iloc[0, [1]].to_string(index=False))< 5.5:
                MU=5
                
                
            #print(merge2[1].iloc[:, [1]].to_string(index=False))
            mu_to_pd = merge2[1].iloc[:, [1]].to_string(index=False).split()
            #print(mu_to_pd[1:])
            energy_to_pd = merge2[1].iloc[:, [2]].to_string(index=False).split()
            angle_to_pd = merge2[1].iloc[:, [0]].to_string(index=False).split()
            
            final['Angle']= [Angle]*36
            final['Mu']= [MU]*36
            final['Energy']= [Energy]*36
            
            final['Moyenne_X']= df_stat_Xmean.mean(axis=1)
            final['Moyenne_Y']= df_stat_Ymean.mean(axis=1)
            
            final['Std_X']= df_stat_Xmean.std(axis=1)
            final['Std_Y']= df_stat_Ymean.std(axis=1)
            
            final['Nominal_X']= X_nominal
            final['Nominal_Y']= Y_nominal
            
            final['True_Angle']= angle_to_pd[1:]
            final['True_Energy']= energy_to_pd[1:]
            final['True_Mu']= mu_to_pd[1:]
            
            
            if creat>0:
                book = load_workbook('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New/Log.xlsx')
                writer = pd.ExcelWriter('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New/Log.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                
                for sheetname in writer.sheets:
                    final.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
                
                writer.save()
        
            if creat==0:
                final.to_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/New/Log.xlsx',index=False)
                creat=creat+1
                
             
            
            merge= []
            merge2=[]
            print('et de 1')
        

        else : pass #print('pas encore 5')
        
        
# =============================================================================
# 
# df = pd.read_csv("/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/Log/log_0096.csv", delimiter=',')#, skiprows=[0])
# 
# #print(df.iloc[:, [3,5,6,7,8]].to_string(index=False))
# 
# df2 = df[['gantry_angle', 'mu', 'energy', 'x','y']]
# 
# 
# df2 = df2.rename(columns={'gantry_angle': 'Angle', 'mu': 'Mu', 'energy': 'Energy', 'x': 'X','y':'Y'})
# df2 = df2.iloc[::-1]
# 
# df2.sort_values(['Y','X'],
#                            ascending=[False,False], inplace=True,kind='mergesort')
# print(df2)
# df2.to_excel('/Users/thomasstinglhamber/Desktop/PHYS22M/Mémoire/Groningen/output.xlsx', index=False)
# 
# 
# =============================================================================









